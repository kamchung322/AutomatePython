# https://openpyxl.readthedocs.org/.
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

def getSampleSheet():
    wb = openpyxl.load_workbook('example.xlsx')
    sht1 = wb['Sheet1']
    return sht1

def createWorkbook():
    wb = openpyxl.Workbook()
    sht2 = wb.create_sheet('NiceSheet')
    sht2['A1'].value = "I'm A1"
    wb.save('Ch13-Sample.xlsx')
    wb.close()

def listWorkSheet():
    wb = openpyxl.load_workbook('example.xlsx')
    print(wb.sheetnames)
    sht1 = wb['Sheet1']
    print(sht1.title)

def singleCell():
    sht1 = getSampleSheet()
    cell = sht1['B1']
    print('Row %s, Column %s is %s' % (cell.row, cell.column, cell.value))
    print("B1 is : %s" %(sht1.cell(row=1, column=2).value))
    print("Max Row %s, Max Col %s" %(sht1.max_row, sht1.max_column))
    print("Max column in English %s" %(get_column_letter(sht1.max_column)))
    print("Eng to Index %s" %(column_index_from_string("AA")))

def listCellRange():
    sht1 = getSampleSheet()

    for row in sht1['A1':'C3']:
        for cell in row:
            print(cell.coordinate, cell.value)
    
    print("*** END OF ROW ***")

def listCellByColumn():
    sht1 = getSampleSheet()
    for cell in list(sht1.columns)[1]:
        print(cell.value)

createWorkbook()
#listCellByColumn()
#listCellRange()
#singleCell()
#listWorkSheet()

